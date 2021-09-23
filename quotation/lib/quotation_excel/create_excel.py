import openpyxl
from django.conf import settings
from quotation.models import Quotations, Quotations_details, Clients
import datetime
from quotation.lib import calculation_module


def create_excel(request, pk):
    Quotations_details_max_row = 25

    template_path = settings.BASE_DIR + \
        '/quotation/lib/quotation_excel/template.xlsx'.replace('/', '\\')
    created_path = settings.BASE_DIR + \
        '/quotation/lib/quotation_excel/created_excel/'.replace('/', '\\')

    book = openpyxl.load_workbook(template_path, data_only=True)
    sheet = book['Sheet1']

    quotation_resultset = Quotations.objects.get(pk=pk)

    updated_datetime = sheet['F1']
    updated_datetime.value = '発行日 ' + \
        str(datetime.date.today().strftime('%Y/%m/%d'))

    quotation_id = sheet['F2']
    quotation_id.value = 'No ' + str(quotation_resultset.quotation_id)

    client_name = sheet['B6']  # clientsテーブルのname
    client_name.value = str(quotation_resultset.client_id or '')

    title = sheet['C9']
    title.value = str(quotation_resultset.title or '')

    delivery_time = sheet['C10']
    delivery_time.value = str(quotation_resultset.delivery_time or '')

    delivery_location = sheet['C11']
    delivery_location.value = str(quotation_resultset.delivery_location or '')

    payment_condition = sheet['C12']
    payment_condition.value = str(quotation_resultset.payment_condition or '')

    expiry = sheet['C13']
    expiry.value = str(quotation_resultset.expiry or '')

    username = sheet['F14']
    username.value = '担当：' + str(request.user.signature or '')

    Quotations_details_resultset = Quotations_details.objects.filter(
        quotation_id=pk)

    i = 0
    unit_amount_list = []
    for result in Quotations_details_resultset:
        unit_amount_list.append(result.sales_unit_price * result.quantity)
        if i <= Quotations_details_max_row:
            item_id = i + 1
            cell = sheet.cell(row=i + 21, column=2)
            cell.value = item_id

            merchandise = result.merchandise
            cell = sheet.cell(row=i + 21, column=3)
            cell.value = str(merchandise or '')

            quantity = result.quantity
            cell = sheet.cell(row=i + 21, column=4)
            cell.value = quantity

            sales_unit_price = result.sales_unit_price
            cell = sheet.cell(row=i + 21, column=5)
            cell.value = sales_unit_price

            amount = unit_amount_list[i]
            cell = sheet.cell(row=i + 21, column=6)
            cell.value = amount

            i += 1

    subtotal = sheet['F46']
    subtotal.value = calculation_module.sub_total(unit_amount_list)

    consumption_tax = sheet['F47']
    consumption_tax.value = quotation_resultset.consumption_tax

    total_amount = sheet['C17']
    total_amount.value = calculation_module.total_amount(
        unit_amount_list)  # subtotal_value + consumption_tax.value
    total_price = sheet['F48']
    total_price.value = calculation_module.total_amount(unit_amount_list)

    remark = sheet['B49']
    remark.alignment = openpyxl.styles.Alignment(wrapText=True, vertical='top')
    remark.value = '<備考>\n' + str(quotation_resultset.remark)

    book.save(created_path + quotation_id.value + '.xlsx')

    return {'book': book, 'book_name': quotation_id.value + '.xlsx',
            'book_path': created_path + quotation_id.value + '.xlsx'}
